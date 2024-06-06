# loops through every ALVRS file and creates a dataframe where each row has
# the file name, the tab name, and the contents of the header row only
# (note the subtle inconsistencies!)

import os
import pandas as pd
from openpyxl import load_workbook
import xlrd

# Function to unmerge cells in each worksheet for .xlsx files
def unmerge_cells_xlsx(sheet):
    for merged_cell in sheet.merged_cells.ranges:
        sheet.unmerge_cells(str(merged_cell))

# Function to unmerge cells in each worksheet for .xls files
def unmerge_cells_xls(sheet):
    for crange in sheet.merged_cells:
        rlo, rhi, clo, chi = crange
        assert rlo < rhi
        assert clo < chi
        for rowx in range(rlo, rhi):
            for colx in range(clo, chi):
                if (rowx, colx) != (rlo, clo):
                    sheet.write(rowx, colx, sheet.cell_value(rlo, clo))

# Function to read rows above the row starting with "AUTAUGA" in each sheet of an Excel file
def read_autauga_rows(file_path):
    _, file_extension = os.path.splitext(file_path)
    if file_extension == '.xlsx':
        wb = load_workbook(filename=file_path)
        sheet_names = wb.sheetnames
        data = []
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            unmerge_cells_xlsx(sheet)  # Unmerge cells in the sheet
            autauga_row_index = None
            for row_idx, row in enumerate(sheet.iter_rows()):
                if row[0].value == "AUTAUGA":
                    autauga_row_index = row_idx
                    break
            if autauga_row_index is not None:
                for row in sheet.iter_rows(min_row=1, max_row=autauga_row_index - 1):
                    autauga_rows = [cell.value for cell in row]
                    data.append((os.path.basename(file_path), sheet_name, autauga_rows))
        return data
    elif file_extension == '.xls':
        wb = xlrd.open_workbook(file_path)
        sheet_names = wb.sheet_names()
        data = []
        for sheet_name in sheet_names:
            sheet = wb.sheet_by_name(sheet_name)
            unmerge_cells_xls(sheet)  # Unmerge cells in the sheet
            autauga_row_index = None
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                if row[0] == "AUTAUGA":
                    autauga_row_index = row_idx
                    break
            if autauga_row_index is not None:
                for row_idx in range(autauga_row_index):
                    autauga_rows = sheet.row_values(row_idx)
                    data.append((os.path.basename(file_path), sheet_name, autauga_rows))
        return data
    else:
        return []


# Main function to process all Excel files in the folder
def process_excel_files(folder_path):
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xls') or f.endswith('.xlsx')]
    all_data = []
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        all_data.extend(read_autauga_rows(file_path))
    df = pd.DataFrame(all_data, columns=['File', 'Tab', 'Autauga_Row'])
    # Filter out rows not containing "WHITE" (case insensitive)
    df = df[df['Autauga_Row'].astype(str).str.contains('WHITE', case=False)]
    return df

# Path to the folder containing the Excel files
folder_path = "/home/linuxlaptop/Documents/AL_elections/data/zz voter registration"

# Process the Excel files and create the dataframe
df = process_excel_files(folder_path)

# Print the first few rows of the filtered dataframe
print(df.head())

# Export the dataframe to excel
df.to_excel("/home/linuxlaptop/Documents/AL_elections/data/zz voter registration/combined.xlsx") # added this


